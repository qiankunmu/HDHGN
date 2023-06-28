import torch
from torch_geometric.loader import DataLoader
from sklearn.metrics import accuracy_score
import numpy as np
from openpyxl import load_workbook
import os
import xlsxwriter
from tqdm import tqdm
import sys

sys.path.append("..")

from vocab import Vocab
from MyDataset import HDHGNDataset_java
from models.HDHGN import HDHGN
from utils import show_2scores, show_score


def main():
    v = Vocab.load("../data/vocab4ast_java.json")
    dataset = HDHGNDataset_java("../data/train_java", "../data/train_files_paths_java.txt", v)
    dataloader = DataLoader(dataset, batch_size=32, shuffle=True)
    valid_dataset = HDHGNDataset_java("../data/valid_java", "../data/valid_files_paths_java.txt", v)
    valid_dataloader = DataLoader(valid_dataset, batch_size=256, shuffle=False)
    test_dataset = HDHGNDataset_java("../data/test_java", "../data/test_files_paths_java.txt", v)
    test_dataloader = DataLoader(test_dataset, batch_size=256, shuffle=False)

    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    num_types = len(v.vocab["types"].word2id)
    vocab_sizes = [len(v.vocab[t].word2id) for t in v.vocab["types"].word2id]
    edge_vocab_size = len(v.vocab["edge_types"].word2id)
    embed_size = 128
    dim_size = 128
    num_layers = 4
    num_edge_heads = 8
    num_node_heads = 8
    num_heads = 8
    feed_sizes = [dim_size, 1024, len(v.vocab["labels"].word2id)]
    dropout_rate = 0.2

    model_name = "HDHGN_java"
    parameters_des = "embed size=" + str(embed_size) + " dim size=" + str(dim_size) + " num layers=" + str(
        num_layers) + " num_edge_heads=" + str(num_edge_heads) + "num_node_heads=" + str(
        num_node_heads) + " num_heads=" + str(num_heads) + " dropout=" + str(dropout_rate) + " feed size=" + "".join(
        [str(f) + " " for f in feed_sizes])
    result_save_path = "../work_dir/HDHGN/"
    if not os.path.exists(result_save_path):
        os.makedirs(result_save_path)

    model_save_path = result_save_path + model_name + ".pt"

    model = HDHGN(num_types, vocab_sizes, edge_vocab_size, embed_size, dim_size, num_layers, num_edge_heads,
                  num_node_heads, num_heads, feed_sizes, dropout_rate)
    model = model.to(device)
    loss_function = torch.nn.CrossEntropyLoss()
    optimizer = torch.optim.Adam(model.parameters(), lr=5e-5)
    # Use ReduceLROnPlateau to realize dynamic learning rate reducing based on accuracy of valid dataset can make result better.
    # However, previous work didn't use the dynamic learning rate, to ensure the comparison of our model's result with previous work results
    # in paper is fair, we fix our learning rate to 5e-5.
    # If you want to use the dynamic learning rate, you can remove the following annotations and annotation in line 88. This will make the
    # last result better and a little higher than results in our paper.
    # scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode="max", factor=0.1, patience=1, verbose=True,
    #                                                        eps=1e-12)

    num_epochs = 50
    max_attk, attk = 5, 0
    max_accuracy = 0
    m_epoch = 0
    loss_list, valid_loss_list, train_acc_list, valid_acc_list = [], [], [], []
    for epoch in tqdm(range(num_epochs)):
        model.train()
        train_loss = 0
        i = 0
        for i, batch_data in enumerate(dataloader):
            batch_data = batch_data.to(device)
            output = model(batch_data.x, batch_data.types, batch_data.edge_types, batch_data.edge_in_out_indexs,
                           batch_data.edge_in_out_head_tail, batch_data.batch)
            loss = loss_function(output, batch_data.labels)
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            train_loss += loss.item()
            if (i + 1) % 200 == 0:
                print(f"epoch={epoch} loss={train_loss / (i + 1)}")

        train_loss /= (i + 1)
        valid_loss, valid_accuracy = valid(model, valid_dataloader, device)
        # scheduler.step(valid_accuracy)

        loss_list.append(train_loss)
        valid_loss_list.append(valid_loss)
        valid_acc_list.append(valid_accuracy)

        print(
            f"epoch finish train loss={train_loss} valid loss={valid_loss} valid accuracy={valid_accuracy}")

        if valid_accuracy > max_accuracy:
            torch.save(model, model_save_path)
            print("successfully save")
            max_accuracy = valid_accuracy
            m_epoch = epoch
            attk = 0
        else:
            attk += 1
            print("no better than last time")
        if attk >= max_attk:
            model = torch.load(model_save_path)
            model = model.to(device)
            print("reload last model")
            attk = 0

    model = torch.load(model_save_path)
    model = model.to(device)
    test_loss, test_accuracy = valid(model, test_dataloader, device)

    print(
        f"train finish max valid accuracy={max_accuracy} 对应epoch={m_epoch} test accuracy={test_accuracy} test loss={test_loss}")
    show_2scores(loss_list, valid_loss_list, "Loss", "train", "valid", "blue", "red",
                 result_save_path + model_name + "-loss.png")
    show_score(valid_acc_list, "Accuracy", "valid", "red", result_save_path + model_name + "-accuracy.png")

    if not os.path.exists("../work_dir/results.xlsx"):
        workbook = xlsxwriter.Workbook("../work_dir/results.xlsx")
        worksheet1 = workbook.add_worksheet("sheet1")
        worksheet1.activate()
        title = ['model', 'parameter', 'valid max accuracy', 'test accuracy']
        worksheet1.write_row('A1', title)
        workbook.close()
    workbook = load_workbook("../work_dir/results.xlsx")
    worksheet = workbook.active
    worksheet.append([model_name, parameters_des, max_accuracy, test_accuracy])
    workbook.save("../work_dir/results.xlsx")
    print("finish")


def valid(model, dataloader, device):
    model.eval()
    loss_function = torch.nn.CrossEntropyLoss()

    losses = []
    preds, labels = [], []
    with torch.no_grad():
        for i, batch_data in enumerate(dataloader):
            batch_data = batch_data.to(device)
            output = model(batch_data.x, batch_data.types, batch_data.edge_types, batch_data.edge_in_out_indexs,
                           batch_data.edge_in_out_head_tail, batch_data.batch)
            loss = loss_function(output, batch_data.labels)
            losses.append(loss.item())
            pred = torch.argmax(output, dim=-1)
            preds.extend(pred.cpu().detach().numpy().tolist())
            labels.extend(batch_data.labels.cpu().detach().numpy().tolist())
    avg_loss = np.mean(losses)
    accuracy = accuracy_score(labels, preds)

    return avg_loss, accuracy


if __name__ == '__main__':
    print("start training on java250")
    main()
